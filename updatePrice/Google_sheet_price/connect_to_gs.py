import json
import logging
import os
import re

import colorlog
import ezsheets
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

handler = colorlog.StreamHandler()
handler.setFormatter(colorlog.ColoredFormatter(
    "%(log_color)s%(levelname)s:%(name)s:%(message)s",
    log_colors={
        'DEBUG': 'cyan',
        'INFO': 'green',
        'WARNING': 'yellow',
        'ERROR': 'red',
        'CRITICAL': 'bold_red',
    }
))

logger = colorlog.getLogger('my_logger')
logger.addHandler(handler)
logger.setLevel(logging.DEBUG)


class GPrice:

    def __init__(self,
                 sheet_id,
                 path_file=None,
                 item_number_google_sheets='B',
                 title_google_sheets='C',
                 price_column_google_sheets='E',
                 recommended_retail_price_column_google_sheets='G',
                 amount_column_google_sheets='H',
                 available_column_google_sheets='I',
                 item_number_export_file='B',
                 price_column_export_file='H',
                 available_column_export_file='G',
                 amount_column_export_file='F',

                 ):
        self.sheet_id = sheet_id
        self.path_file = path_file
        self.item_number_google_sheets = item_number_google_sheets
        self.title_google_sheets = title_google_sheets
        self.price_column_google_sheets = price_column_google_sheets
        self.recommended_retail_price_column_google_sheets = recommended_retail_price_column_google_sheets
        self.available_column_google_sheets = available_column_google_sheets
        self.amount_column_google_sheets = amount_column_google_sheets
        self.item_number_export_file = item_number_export_file
        self.price_column_export_file = price_column_export_file
        self.available_column_export_file = available_column_export_file
        self.amount_column_export_file = amount_column_export_file
        self.items_sheet_list = []
        self.items_excel_list = []

    @staticmethod
    def _not_valid_item_number(vc):
        title_list = ['ФОТО', 'Артикул', 'Наименование товара', 'Основные параметры',
                      'Цена/USD', 'Цена/UAH', 'Количество в ящике', 'Наличие', 'Цена']

        if vc == '':
            return True
        elif vc is None:
            return True

        for name in title_list:
            if re.match(fr'^{name}', vc):
                return True

    @staticmethod
    def _replaceSpace(cell_value):
        return ' '.join(cell_value.split())

    @staticmethod
    def _valid_price(price):

        symbols = ['$', ',']

        if isinstance(price, str):
            for symbol in symbols:
                price = price.replace(symbol, '')

        try:
            return float(price)
        except Exception as ex:
            logger.error(f"Not valid price {price}, ERROR: {ex}")

    @staticmethod
    def _valid_availability_export_file(cell):

        if cell == '+':
            return 'TRUE'
        else:
            return 'FALSE'

    def connect_to_google(self):
        try:
            logger.info("Connecting to google sheets...")
            ss = ezsheets.Spreadsheet(self.sheet_id)
            sheet = ss.sheets[0]

            return sheet
        except Exception as ex:
            logger.error("Problem with connecting to google sheets...")
            print(ex)

    def _connect_to_excel(self):
        try:
            logger.info("Connecting to excel sheets...")
            ee = load_workbook(self.path_file)
            excel_sheet = ee.active
            return excel_sheet

        except Exception as ex:
            logger.error("Problem with connecting to excel sheets...")
            print(ex)

    def _get_data_from_excel_file_for_update_price(self):
        excel_sheet = self._connect_to_excel()

        excel_items = {}

        for i in range(1, excel_sheet.max_row + 1):
            item_number = excel_sheet[f'{self.item_number_export_file}{i}'].value

            if self._not_valid_item_number(item_number):
                continue

            price_excel = self._valid_price(excel_sheet[f"{self.price_column_export_file}{i}"].value)

            if price_excel is None:
                continue

            amount = excel_sheet[f"{self.amount_column_export_file}{i}"].value

            availability_excel = self._valid_availability_export_file(
                excel_sheet[f"{self.available_column_export_file}{i}"].value)

            excel_items[item_number] = {
                'price': price_excel,
                'amount': amount,
                'availability': availability_excel
            }

        if not excel_items:
            logger.warning("The Excel file does not contain valid data.")
        return excel_items

    def _set_data_to_google_sheets_for_update_price(self, excel_items):

        if not excel_items:
            logger.warning("No data to update in Google Sheets.")
            return

        sheet = self.connect_to_google()

        for i in range(1, sheet.rowCount + 1):
            item_number = sheet[f'{self.item_number_google_sheets}{i}']

            if self._not_valid_item_number(item_number):
                continue

            title = self._replaceSpace(sheet[f'{self.title_google_sheets}{i}'])
            sheet[f'{self.title_google_sheets}{i}'] = title

            if item_number in excel_items:
                data = excel_items[item_number]

                sheet[f'{self.amount_column_google_sheets}{i}'] = data['amount']
                sheet[f'{self.price_column_google_sheets}{i}'] = data['price']
                sheet[f'{self.available_column_google_sheets}{i}'] = data['availability']
                logger.info(f"Product data has been updated {item_number}.")

            else:
                sheet[f'{self.available_column_google_sheets}{i}'] = "FALSE"
                logger.info(f"The item {item_number} was not found in Excel. Available set 'False'.")

    def _get_data_from_excel_file(self):
        logger.info("Getting data from an Excel file...")

        excel_sheet = self._connect_to_excel()

        excel_items = {}

        for i in range(2, excel_sheet.max_row + 1):
            item_number = excel_sheet[f'{self.item_number_export_file}{i}'].value

            if self._not_valid_item_number(item_number):
                continue

            rrp_price_excel = self._valid_price(excel_sheet[f"{self.price_column_export_file}{i}"].value)

            if rrp_price_excel is None:
                continue

            amount = excel_sheet[f"{self.amount_column_export_file}{i}"].value

            excel_items[item_number] = {
                'price': rrp_price_excel,
                'amount': amount
            }

        if not excel_items:
            logger.warning("The Excel file does not contain valid data.")
        return excel_items

    def _set_data_to_google_sheets(self, excel_items):

        if not excel_items:
            logger.warning("No data to update in Google Sheets.")
            return

        sheet = self.connect_to_google()

        for i in range(1, sheet.rowCount + 1):
            item_number = sheet[f'{self.item_number_google_sheets}{i}']

            if self._not_valid_item_number(item_number):
                continue

            if item_number in excel_items:
                data = excel_items[item_number]

                sheet[f'{self.amount_column_google_sheets}{i}'] = data['amount']
                sheet[f'{self.recommended_retail_price_column_google_sheets}{i}'] = data['price']
                logger.info(f"Product data has been updated {item_number}.")
            else:
                sheet[f'{self.recommended_retail_price_column_google_sheets}{i}'] = 0
                logger.info(f"The item {item_number} was not found in Excel. The price is set to 0.")

    def updatePrice(self):

        excel_items = self._get_data_from_excel_file_for_update_price()

        self._set_data_to_google_sheets_for_update_price(excel_items)

    def updateRRP(self):

        excel_items = self._get_data_from_excel_file()

        self._set_data_to_google_sheets(excel_items)


def main():

    google_sheets_price = 'GRAND_ELTOS'
    excel_price = r"C:\Users\user\Downloads\Прайс Grand 18 вересня 2024_2 (1).xlsx"
    excel_price_RRP = r"C:\Users\user\Downloads\Telegram Desktop\Прайс_Grand_з_рекомендованою_мінімальною_роздрібною_ціною_на_вересень.xlsx"

    up = GPrice(
        sheet_id=os.getenv(google_sheets_price),
        path_file=excel_price,
    )

    # excel_price
    up.updatePrice()

    # # excel_price_RRP
    # up.updateRRP()


if __name__ == '__main__':
    main()
