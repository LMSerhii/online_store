import datetime
import json
import math
import os
import re
import time

import ezsheets
from dotenv import load_dotenv
from openpyxl import load_workbook

from paths import epic_press, roz_press, prom_press, prom_grand, roz_grand, epic_grand

load_dotenv()


class UpdatePrice:
    def __init__(self, export_path="", marg=100, or_marg=400, curr=40, rate_sell=15, vcc="D", valuta='USD'):
        self.export_path = export_path
        self.margin = marg
        self.or_margin = or_marg
        self.curr = curr
        self.rate_sell = rate_sell
        self.vendor_code_coll = vcc
        self.valuta = valuta

    def royalty(self, price, rate):
        result = (price / (100 - rate)) * 100
        return math.ceil(result)

    def __vendor_code(self, worksheet, itr, rate):

        vendor_code = worksheet[f'{self.vendor_code_coll}{itr}'].value

        if re.match(r'^OR\|+', vendor_code):

            if '||' in vendor_code:

                pp = vendor_code.split('||')[-1]

                new_price = self.royalty((float(pp) * self.curr + self.or_margin), rate)

                old_price = self.royalty(new_price, self.rate_sell)

                return new_price, old_price

            elif '|' in vendor_code:

                pp = vendor_code.split('|')[-1]

                if re.search(r'Т', vendor_code):
                    new_price = self.royalty(
                        (float(pp) + self.or_margin + 150), rate)
                else:
                    new_price = self.royalty(
                        (float(pp) + self.or_margin), rate)

                old_price = self.royalty(new_price, self.rate_sell)

                return new_price, old_price

        else:

            if '||' in vendor_code:

                pp = vendor_code.split('||')[-1]

                new_price = self.royalty(
                    (float(pp) * self.curr + self.margin), rate)

                old_price = self.royalty(new_price, self.rate_sell)

                return new_price, old_price

            else:

                pp = vendor_code.split('|')[-1]

                if re.search(r'Т', vendor_code):
                    new_price = self.royalty(
                        (float(pp) + self.margin + 150), rate)
                else:
                    new_price = self.royalty((float(pp) + self.margin), rate)

                old_price = self.royalty(new_price, self.rate_sell)

                return new_price, old_price

    def __availability(self, worksheet, itr):

        availability = worksheet[f'P{itr}'].value

        if availability == '+' or availability == '!':
            discount = f'{self.rate_sell}%'

            date_start = datetime.datetime.fromtimestamp(time.time()).strftime('%d.%m.%Y')

            date_end = datetime.datetime.fromtimestamp(time.time() + 604800).strftime('%d.%m.%Y')

            return discount, date_start, date_end

        else:
            discount = ''
            date_start = ''
            date_end = ''

            return discount, date_start, date_end

    @staticmethod
    def __get_rate(export_rate_id):
        try:
            with open('prom_rate.json', 'r', encoding='utf-8') as f:
                file = json.load(f)

            for item in file:
                if 'cat_id' in item and 'rate' in item:
                    prom_rate_id = int(item.get('cat_id'))

                    if export_rate_id == prom_rate_id:
                        rate = item.get('rate')
                        if rate.endswith('%'):
                            return float(rate[:-1])
            return 0

        except Exception as ex:
            print('=' * 50)
            print(ex)
            print('-' * 50)
            print('export_rate_id', export_rate_id)
            print('=' * 50)

    @staticmethod
    def __vendor_validation(vc_export):

        if vc_export is None:
            return False

        if re.match(r'^Код_товара', vc_export):
            return False

        return True

    def __get_price(self, sheet_id, vc_export):
        ss = ezsheets.Spreadsheet(sheet_id)
        sheet = ss.sheets[0]

        for i in range(1, sheet.rowCount + 1):
            vendor_code = sheet[f'B{i}']
            if vendor_code == '':
                continue

            if re.search(fr"{vendor_code}\|\w+|{vendor_code}\|\|\w+|^{vendor_code}$", vc_export):
                print(f'{vendor_code} == {vc_export}')

                if self.valuta == 'USD':
                    price = sheet[f'E{i}']
                elif self.valuta == 'UAH':
                    price = sheet[f'F{i}']
                else:
                    price = 0

                if isinstance(price, str) and '$' in price:
                    price = price.replace('$', '').replace(',', '.').strip()

                availability = sheet[f'H{i}']

                return float(price), availability, vendor_code

    def __put_id_prom(self, worksheet, itr, sheet_id):

        vencod_export = worksheet[f'{self.vendor_code_coll}{itr}'].value

        result = self.__get_price(sheet_id=sheet_id, vc_export=vencod_export)

        if result is not None:

            price, availability, vendor_code = result

            if availability == "TRUE":
                worksheet[f'P{itr}'].value = '!'
            elif availability == "FALSE":
                worksheet[f'P{itr}'].value = '-'

            if re.search(r"\|\|", vencod_export):

                if re.search(fr"\bOR", vencod_export):
                    worksheet[f'{self.vendor_code_coll}{itr}'].value = f'OR|{vendor_code}||000{price}'
                else:
                    worksheet[f'{self.vendor_code_coll}{itr}'].value = f'{vendor_code}||000{price}'

            else:

                if re.search(fr"\bOR", vencod_export):
                    worksheet[f'{self.vendor_code_coll}{itr}'].value = f'OR|{vendor_code}|000{price}'
                else:
                    worksheet[f'{self.vendor_code_coll}{itr}'].value = f'{vendor_code}|000{price}'

            return True
        else:
            return False

    def updateProm(self, rate_column='AA', from_price=None):
        wb = load_workbook(filename=self.export_path)
        ws = wb.active

        for i in range(2, ws.max_row + 1):

            vc_export = ws[f'{self.vendor_code_coll}{i}'].value

            if not self.__vendor_validation(vc_export):
                continue

            export_rate_id = ws[f'{rate_column}{i}'].value

            rate = self.__get_rate(export_rate_id)

            if from_price is not None:
                if not self.__put_id_prom(ws, i, from_price):
                    print(f"{ws[f'A{i}'].value} was not found")

            new_price, old_price = self.__vendor_code(
                worksheet=ws, itr=i, rate=rate)

            discount, date_start, date_end = self.__availability(
                worksheet=ws, itr=i)

            ws[f'I{i}'].value = str(old_price)
            ws[f'AE{i}'].value = discount

            ws[f'AI{i}'].value = date_start
            ws[f'AJ{i}'].value = date_end

            print('=' * 70)

        wb.save(self.export_path)

        return "Successful updated"

    def updateEpik(self, rate, from_price=None):
        wb = load_workbook(filename=self.export_path)
        ws = wb.active

        for i in range(2, ws.max_row + 1):

            if from_price is not None:

                vencod_export = ws[f'{self.vendor_code_coll}{i}'].value

                result = self.__get_price(
                    sheet_id=from_price, vc_export=vencod_export)

                if result is not None:

                    price, availability, vendor_code = result

                    if availability == "TRUE":
                        ws[f'H{i}'].value = 'в наявності'
                    elif availability == "FALSE":
                        ws[f'H{i}'].value = 'немає в наявності'

                    if re.search(r"\|\|", vencod_export):

                        if re.match(r"^OR\|+", vencod_export):
                            new_price = self.royalty(
                                price * self.curr + self.or_margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)
                        else:
                            new_price = self.royalty(
                                price * self.curr + self.margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)

                    else:

                        if re.match(r"^OR\|+", vencod_export):
                            new_price = self.royalty(
                                price + self.or_margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)
                        else:
                            new_price = self.royalty(
                                price + self.margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)

                    ws[f'E{i}'].value = str(new_price)
                    ws[f'F{i}'].value = str(old_price)

                else:
                    print(f'{vencod_export} was not found')

            else:
                new_price, old_price = self.__vendor_code(
                    worksheet=ws, itr=i, rate=rate)

                ws[f'E{i}'].value = str(new_price)
                ws[f'F{i}'].value = str(old_price)

        wb.save(f'{self.export_path}')

        return "Successful updated"

    def updateRozetka(self, rate, from_price=None):
        wb = load_workbook(filename=self.export_path)
        ws = wb.active

        for i in range(2, ws.max_row + 1):

            if from_price is not None:

                vencod_export = ws[f'{self.vendor_code_coll}{i}'].value

                result = self.__get_price(
                    sheet_id=from_price, vc_export=vencod_export)

                if result is not None:

                    price, availability, vendor_code = result

                    if availability == "TRUE":
                        ws[f'P{i}'].value = 'Есть в наличии'
                    elif availability == "FALSE":
                        ws[f'P{i}'].value = 'Нет в наличии'

                    if re.search(r"\|\|", vencod_export):

                        if re.match(r"^OR\|+", vencod_export):
                            new_price = self.royalty(
                                price * self.curr + self.or_margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)
                        else:
                            new_price = self.royalty(
                                price * self.curr + self.margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)

                        ws[f'I{i}'].value = str(new_price)
                        ws[f'J{i}'].value = str(old_price)

                    else:
                        if re.match(r"^OR\|+", vencod_export):
                            new_price = self.royalty(
                                price + self.or_margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)
                        else:
                            new_price = self.royalty(
                                price + self.margin, rate)
                            old_price = self.royalty(new_price, self.rate_sell)

                        ws[f'I{i}'].value = str(new_price)
                        ws[f'J{i}'].value = str(old_price)
            else:
                new_price, old_price = self.__vendor_code(
                    worksheet=ws, itr=i, rate=rate)

                ws[f'I{i}'].value = new_price
                ws[f'J{i}'].value = old_price

        wb.save(self.export_path)

        return "Successful updated"


def prom(path, prices, margin=70, original_margin=300, current_course=39, rate_sell=20, valuta='USD',
         vendor_code_column='A'):
    price_lists = prices

    export = UpdatePrice(
        export_path=path,
        marg=margin,
        or_marg=original_margin,
        curr=current_course,
        rate_sell=rate_sell,
        vcc=vendor_code_column,
        valuta=valuta
    )

    for price in price_lists:
        print(export.updateProm(from_price=os.getenv(price)))


def epicentr(base_dir, prices, margin=100, original_margin=300, current_course=39, rate_sell=20, valuta='USD',
             vendor_code_column='D'):
    BASE_DIR = base_dir
    PRICE_LISTS = prices

    for path in os.listdir(BASE_DIR):

        path_to_file = os.path.join(BASE_DIR, path)

        export = UpdatePrice(
            export_path=path_to_file,
            marg=margin,
            or_marg=original_margin,
            curr=current_course,
            rate_sell=rate_sell,
            vcc=vendor_code_column,
            valuta=valuta
        )

        with open("epik_rate.json", "r", encoding="utf-8") as f:
            file = json.load(f)

        for item in file.get('rate'):
            if re.match(rf'{item}', path.split('.')[0]):
                print(f"{item}:{file.get('rate').get(item)}")
                rate = file.get('rate').get(item)
                break

        for price in PRICE_LISTS:
            print(export.updateEpik(rate=rate, from_price=os.getenv(price)))


def rozetka(base_dir, prices, margin=100, original_margin=300, current_course=39, rate_sell=20, valuta='USD',
            vendor_code_column='E'):
    BASE_DIR = base_dir
    PRICE_LISTS = prices

    for path in os.listdir(BASE_DIR):

        path_to_file = os.path.join(BASE_DIR, path)

        export = UpdatePrice(
            export_path=path_to_file,
            marg=margin,
            or_marg=original_margin,
            curr=current_course,
            rate_sell=rate_sell,
            vcc=vendor_code_column,
            valuta=valuta
        )

        with open("rozetka_rate.json", "r", encoding="utf-8") as f:
            file = json.load(f)

        for item in file.get('rate'):
            if re.match(rf'{item}', path.split('.')[0]):
                rate = file.get('rate').get(item)
                print(f"{item}:{file.get('rate').get(item)}")
                break

        for price in PRICE_LISTS:
            print(export.updateRozetka(rate=rate, from_price=os.getenv(price)))


def manual(margin, original_margin, rate, rate_sell, curr):
    export = UpdatePrice(curr=curr)
    while True:
        price = float(input('Enter price: '))
        new_price = export.royalty(price + margin, rate)
        old_price = export.royalty(new_price, rate_sell)
        or_new_price = export.royalty(price + original_margin, rate)
        or_old_price = export.royalty(or_new_price, rate_sell)
        print("=== TRUE PRICE ===")
        print(new_price)
        print(old_price)
        print("=== ORIGINAL PRICE ===")
        print(or_new_price)
        print(or_old_price)
        print("++++++ END +++++++")


def main():
    # PRESS, KORM, GAZ, INCUBATOR, KITCHEN, GRAND_ELTOS

    MARKETPLACE = 'PROM'

    match MARKETPLACE:
        case "MANUAL":
            manual(margin=150, original_margin=450, rate=15.15, rate_sell=20, curr=41)

        case "PROM":
            prom(path=prom_grand, prices=["GRAND_ELTOS"], valuta='USD',
                 current_course=41.7, margin=100, original_margin=430, rate_sell=45)
            # prom_press, prom_korm, prom_incubator, prom_gaz, prom_grand

        case "EPICENTR":
            epicentr(base_dir=epic_grand, prices=['GRAND_ELTOS'], valuta='USD', current_course=41.7, margin=250,
                     original_margin=500)
            # epic_press, epic_incubator, epic_korm, epic_gaz, epic_house_tech, epic_grand

        case "ROZETKA":
            rozetka(base_dir=roz_grand, margin=250, original_margin=500,
                    prices=["GRAND_ELTOS"],
                    valuta='USD', current_course=41.7)
            # roz_press, roz_gaz, roz_incubator,roz_grand

        case _:
            print("You do not have any access to the code")


if __name__ == '__main__':
    main()
