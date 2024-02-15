import os
import re
import json

from openpyxl import load_workbook
from tqdm import tqdm

from api_request import sendRequest


class Status:

    def getStatus(self, barcode):

        req = sendRequest()

        if re.match(r'^204|^59', f'{barcode}'):
            response = req.request_to_np(barcode=barcode)
            return response.get('data')[0].get('Status')
        elif re.match(r'^050', f'{barcode}'):
            response = req.request_to_ukr(barcode=barcode)
            return response.get('eventName')
        elif re.match(r'^50', f'{barcode}'):
            response = req.request_to_ukr(barcode=f"0{barcode}")
            return response.get('eventName')
        else:
            return ''

    def readExcel(self, path_to_path, ttn_col='H', status_col='L'):
        wb = load_workbook(filename=path_to_path)
        sheet = wb.active

        barcodeList = []

        for i in tqdm(range(2, sheet.max_row + 1)):

            barcode = sheet[f'{ttn_col}{i}'].value

            if not barcode:
                continue

            barcodeList.append({
                "barcode": barcode,
                "id": sheet[f'A{i}'].value,
            })

        with open('barcodeList.json', 'w', encoding='utf-8') as file:
            json.dump(barcodeList, file, indent=4, ensure_ascii=False)

        for i in tqdm(range(2, sheet.max_row + 1), position=0):
            barcode = sheet[f'{ttn_col}{i}'].value
            sheet[f'{status_col}{i}'].value = self.getStatus(barcode)

        name = os.path.split(path_to_path)[-1]
        wb.save(path_to_path)


def main():
    path_file = r"D:\Works\Online_store\prom_ua_report_status\data\November.xlsx"

    st = Status()
    st.readExcel(path_file, status_col='Q')

    # print(st.getStatus(barcode='0504270497177'))


if __name__ == "__main__":
    main()
