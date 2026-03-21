import datetime
import json
import logging
import math
import os
import re
import pprint
from csv import excel_tab
from numbers import Number
import unicodedata
import colorlog
import ezsheets
from dotenv import load_dotenv
from lxml.html.builder import OBJECT
from openpyxl import load_workbook
from pyasn1.type.univ import Boolean

from paths import *

load_dotenv()
handler = colorlog.StreamHandler()
handler.setFormatter(
    colorlog.ColoredFormatter(
        "%(log_color)s%(levelname)s:%(name)s:%(message)s",
        log_colors={
            "DEBUG": "cyan",
            "INFO": "green",
            "WARNING": "yellow",
            "ERROR": "red",
            "CRITICAL": "bold_red",
        },
    )
)

logger = colorlog.getLogger("my_logger")
logger.addHandler(handler)
logger.setLevel(logging.DEBUG)

logging.getLogger().handlers.clear()

DEFAULT_RATE = 15


class UpdatePrice:
    def __init__(
        self,
        worksheet=None,
        margin: int | float | str = 100,
        original_margin: int | float | str = 400,
        currency_rate=40,
        rate_sell: int | float = 15,
        vendor_code_column="D",
        valuta="USD",
        rrp_column="G",
    ):
        self.worksheet = worksheet
        self.margin = margin
        self.original_margin = original_margin
        self.currency_rate = currency_rate
        self.rate_sell = rate_sell
        self.vendor_code_column = vendor_code_column
        self.valuta = valuta
        self.rrp_column = rrp_column

    @staticmethod
    def royalty(price, rate):
        return math.ceil((price / (100 - rate)) * 100)

    @staticmethod
    def _resolve_margin(margin_spec, base: float) -> float:
        """Абсолют (int/float або рядок числа) або відсоток від `base` (рядок із '%')."""
        if margin_spec is None:
            return 0.0
        if isinstance(margin_spec, str):
            s = margin_spec.strip()
            if s.endswith("%"):
                try:
                    pct = float(s[:-1].strip().replace(",", "."))
                except ValueError:
                    logger.error(f"Invalid margin percent: {margin_spec!r}")
                    return 0.0
                return base * (pct / 100.0)
            try:
                return float(s.replace(",", "."))
            except ValueError:
                logger.error(f"Invalid margin value: {margin_spec!r}")
                return 0.0
        return float(margin_spec)

    @staticmethod
    def _format_prom_discount_pct(pct: float) -> str:
        """Відсоток для колонки знижки: допускає дробову частину, крапка як роздільник."""
        p = float(pct)
        if not math.isfinite(p):
            return "0%"
        p = round(p, 8)
        if abs(p - round(p)) < 1e-12:
            return f"{int(round(p))}%"
        s = f"{p:.8f}".rstrip("0").rstrip(".")
        return f"{s}%"

    @staticmethod
    def _prom_discount_absolute_uah(
        old_price: float,
        nominal_discount_pct: float,
    ) -> int:
        """
        Сума знижки в гривнях для колонки AE: old_price − round(old_price × (100 − p) / 100),
        щоб кінцева ціна (без копійок) була найближчою до номінальної відсоткової знижки p.
        """
        if old_price <= 0:
            return 0
        op = int(round(float(old_price)))
        sale = op * (100 - float(nominal_discount_pct)) / 100.0
        sale_i = int(round(sale))
        delta = op - sale_i
        return max(0, min(op, delta))

    def __vendor_code(self, worksheet, row_number, rate):
        vendor_code = worksheet[f"{self.vendor_code_column}{row_number}"].value

        if vendor_code is None:
            return None, None

        is_original = vendor_code.startswith("OR|")
        has_double_pipe = "||" in vendor_code
        has_T = "Т" in vendor_code

        separator = "||" if has_double_pipe else "|"

        try:
            pp = float(vendor_code.split(separator)[-1])
        except (ValueError, IndexError):
            logger.error(
                f"Error parsing vendor code at row {row_number}: {vendor_code}"
            )
            return None, None

        base_price = pp * self.currency_rate if has_double_pipe else pp

        margin_spec = self.original_margin if is_original else self.margin
        base_price += self._resolve_margin(margin_spec, base_price) + (
            150 if has_T else 0
        )

        new_price = self.royalty(base_price, rate)
        old_price = self.royalty(new_price, self.rate_sell)

        return new_price, old_price

    def __availability(self, worksheet, row_number):

        availability = worksheet[f"P{row_number}"].value

        if availability in ("+", "!"):
            discount = UpdatePrice._format_prom_discount_pct(
                float(self.rate_sell))
            date_start = datetime.datetime.now().strftime("%d.%m.%Y")
            date_end = (datetime.datetime.now() + datetime.timedelta(days=7)).strftime(
                "%d.%m.%Y"
            )
        else:
            discount = ""
            date_start = ""
            date_end = ""

        return discount, date_start, date_end

    _rates_cache = {}

    @staticmethod
    def __load_rates(rate_file="prom_rate.json"):
        if rate_file not in UpdatePrice._rates_cache:
            try:
                with open(rate_file, "r", encoding="utf-8") as f:
                    rates = json.load(f)
                UpdatePrice._rates_cache[rate_file] = {
                    str(item.get("cat_id")): float(item.get("rate", "0%").rstrip("%"))
                    for item in rates
                }
            except Exception as ex:
                logger.error(f"Error reading rate from {rate_file}: {ex}")
                UpdatePrice._rates_cache[rate_file] = {}
        return UpdatePrice._rates_cache[rate_file]

    @staticmethod
    def __load_prom_rate(export_rate_id, rate_file="prom_rate.json"):
        rates = UpdatePrice.__load_rates(rate_file)
        return rates.get(str(export_rate_id), 0.0)

    @staticmethod
    def __vendor_validation(vc_export):

        if vc_export is None:
            return False

        if re.match(r"^Код_товара", vc_export):
            return False

        return True

    _sheet_cache = {}

    @staticmethod
    def __load_sheet(sheet_id):
        if sheet_id not in UpdatePrice._sheet_cache:
            try:
                ss = ezsheets.Spreadsheet(sheet_id)
                sheet = ss[0]
                data = sheet.getRows()
                UpdatePrice._sheet_cache[sheet_id] = {
                    row[1]: row for row in data}
            except Exception as e:
                logger.error(f"Error loading sheet {sheet_id}: {e}")
                UpdatePrice._sheet_cache[sheet_id] = {}
        return UpdatePrice._sheet_cache[sheet_id]

    @staticmethod
    def __extract_vendor_key(vendor_code_export):
        if "OR|" in vendor_code_export:
            return vendor_code_export.split("|")[1]
        return vendor_code_export.split("|")[0]

    def __get_price(self, sheet_id, vendor_code_export):
        vendor_code_dict = self.__load_sheet(sheet_id)

        vendor_code_key = self.__extract_vendor_key(vendor_code_export)

        if vendor_code_key not in vendor_code_dict:
            return None

        row = vendor_code_dict[vendor_code_key]
        price = row[4] if self.valuta == "USD" else row[5]
        availability = row[8]

        price = price.replace(',', '.')
        cleaned = re.sub(r'[^\d.\-]', '', price)

        return float(cleaned), availability, vendor_code_key

    @staticmethod
    def __get_rrp(sheet_id, vendor_code_export):
        vendor_code_dict = UpdatePrice.__load_sheet(sheet_id)

        vendor_code_key = UpdatePrice.__extract_vendor_key(vendor_code_export)

        if vendor_code_key not in vendor_code_dict:
            logger.warning(
                f"Vendor code {vendor_code_export} not found in RRP sheet.")
            return None

        row = vendor_code_dict[vendor_code_key]
        rrp = row[6]
        return float(rrp) if rrp != "" else None

    def __compare_and_set_price(self, worksheet, row_number, sheet_id, final_price):
        rrp = self.__get_rrp(
            sheet_id, worksheet[f"{self.vendor_code_column}{row_number}"].value
        )
        if rrp is not None and final_price < rrp:
            final_price = rrp
        return final_price

    def __put_id_prom(self, worksheet, row_number, sheet_id):
        vendor_code_export = worksheet[f"{self.vendor_code_column}{row_number}"].value

        result = self.__get_price(
            sheet_id=sheet_id, vendor_code_export=vendor_code_export
        )

        if result:
            price, availability, vendor_code = result

            worksheet[f"P{row_number}"].value = "!" if availability == "TRUE" else "-"

            separator = "||" if "||" in vendor_code_export else "|"

            prefix = "OR|" if vendor_code_export.startswith("OR|") else ""
            worksheet[f"{self.vendor_code_column}{row_number}"].value = (
                f"{prefix}{vendor_code}{separator}00{price}"
            )
            return True
        else:
            return False

    @staticmethod
    def resolve_rate(rate, price=None):
        if isinstance(rate, (int, float)):
            return rate

        if isinstance(rate, list):
            for tier in rate:
                min_val = tier.get("min", 0)
                max_val = tier.get("max", float("inf"))

                if price is not None:
                    if min_val <= price < max_val:
                        return tier["rate"]
                else:
                    return tier["rate"]

            return DEFAULT_RATE

        return DEFAULT_RATE

    def updateProm(
        self,
        rate_column="AA",
        from_price=None,
        prom_ae_absolute_uah=True,
    ):
        ws = self.worksheet

        for i in range(2, ws.max_row + 1):
            vendor_code_export = ws[f"{self.vendor_code_column}{i}"].value

            if not self.__vendor_validation(vendor_code_export):
                continue

            export_rate_id = ws[f"{rate_column}{i}"].value
            rate = self.__load_prom_rate(export_rate_id)

            is_found = self.__put_id_prom(ws, i, from_price)

            if from_price and not is_found:
                logger.warning(
                    f"Product at row {i - 1} / {ws.max_row - 1} was not found.")
                continue

            new_price, old_price = self.__vendor_code(ws, i, rate)

            if new_price is None:
                continue

            final_price = self.__compare_and_set_price(
                ws, i, from_price, new_price)

            old_price = self.royalty(final_price, self.rate_sell)

            discount, date_start, date_end = self.__availability(ws, i)
            if discount and prom_ae_absolute_uah:
                ae_value = self._prom_discount_absolute_uah(
                    float(old_price), float(self.rate_sell)
                )
            else:
                ae_value = discount

            ws[f"I{i}"].value = old_price
            ws[f"AE{i}"].value = ae_value
            ws[f"AI{i}"].value = date_start
            ws[f"AJ{i}"].value = date_end

            logger.info(f"Row {i - 1} / {ws.max_row - 1} updated.")

    def updateEpik(self, rate, from_price=None):
        ws = self.worksheet

        for i in range(2, ws.max_row + 1):

            if from_price is not None:
                vencod_export = ws[f"{self.vendor_code_column}{i}"].value

                result = self.__get_price(
                    sheet_id=from_price, vendor_code_export=vencod_export
                )

                if result is None:
                    logger.warning(f"Row {i} / {ws.max_row} completed")
                    logger.warning(f"{vencod_export} was not found")
                    continue

                price, availability, vendor_code = result

                if availability == "TRUE":
                    ws[f"H{i}"].value = "в наявності"
                elif availability == "FALSE":
                    ws[f"H{i}"].value = "немає в наявності"

                has_double = bool(re.search(r"\|\|", vencod_export))
                is_or = bool(re.match(r"^OR\|+", vencod_export))
                base = price * self.currency_rate if has_double else price
                margin_spec = self.original_margin if is_or else self.margin
                addon = self._resolve_margin(margin_spec, base)
                new_price = self.royalty(base + addon, rate)
                old_price = self.royalty(new_price, self.rate_sell)

                ws[f"E{i}"].value = new_price
                ws[f"F{i}"].value = old_price
                logger.info(f"Row {i} / {ws.max_row} completed")

            else:
                new_price, old_price = self.__vendor_code(
                    worksheet=ws, row_number=i, rate=rate
                )

                ws[f"E{i}"].value = new_price
                ws[f"F{i}"].value = old_price
                logger.info(f"Row {i} / {ws.max_row} completed")

    def updateRozetka(self, rate: int | float | list, from_price: str):
        ws = self.worksheet

        for i in range(2, ws.max_row + 1):
            vencod_export = ws[f"{self.vendor_code_column}{i}"].value
            new_price = 0
            old_price = 0
            new_vendor_code = vencod_export

            result = self.__get_price(
                sheet_id=from_price, vendor_code_export=vencod_export
            )

            if result is None:
                logger.warning(f"Row {i - 1} / {ws.max_row - 1} completed")
                logger.warning(f"Vendor code {vencod_export} not found.")
                continue

            price, availability, vendor_code = result

            if availability == "TRUE":
                ws[f"Q{i}"].value = "В наявності"
            elif availability == "FALSE":
                ws[f"Q{i}"].value = "Не в наявності"

            has_double = bool(re.search(r"\|\|", vencod_export))
            is_or = bool(re.match(r"^OR\|+", vencod_export))
            base = price * self.currency_rate if has_double else price
            margin_spec = self.original_margin if is_or else self.margin
            addon = self._resolve_margin(margin_spec, base)
            base_with_margin = base + addon

            selected_rate = self.resolve_rate(rate, price=base_with_margin) * 1.08
            new_price = self.royalty(base_with_margin, selected_rate)
            old_price = self.royalty(new_price, self.rate_sell)

            if new_price == 0 and old_price == 0:
                ws[f"Q{i}"].value = "Не в наявності"
                logger.info(f"Something wrong with price")
                logger.info(f"Rate: {selected_rate} -- Price: {price} -- Finish price: {new_price}")

            divider = "||" if "||" in vencod_export else "|"

            def clean_number(num: float) -> str:
                s = str(num)
                s = s.rstrip('0').rstrip('.')
                return s

            price = clean_number(price)

            if "OR|" in vencod_export:
                new_vendor_code = f"OR|{vendor_code}{divider}000{price}"
            else:
                new_vendor_code = f"{vendor_code}{divider}000{price}"

            ws[f"I{i}"].value = new_price
            ws[f"J{i}"].value = old_price
            ws[f"E{i}"].value = str(new_vendor_code)

            logger.info(f"Row {i - 1} / {ws.max_row - 1} completed")


def prom(
    export_file_path,
    prices,
    margin: int | float | str = 100,
    original_margin: int | float | str = 300,
    current_course: int | float = 39,
    rate_sell: int | float = 20,
    valuta: str = "USD",
    vendor_code_column: str = "A",
):
    wb = load_workbook(filename=export_file_path)
    ws = wb.active

    export = UpdatePrice(
        worksheet=ws,
        margin=margin,
        original_margin=original_margin,
        currency_rate=current_course,
        rate_sell=rate_sell,
        vendor_code_column=vendor_code_column,
        valuta=valuta,
    )

    for price in prices:
        export.updateProm(from_price=os.getenv(price))

    wb.save(filename=export_file_path)


def epicentr(
    base_dir,
    prices,
    margin: int | float | str = 100,
    original_margin: int | float | str = 300,
    current_course: int | float = 39,
    rate_sell: int | float = 20,
    valuta: str = "USD",
    vendor_code_column: str = "D",
):
    for path in os.listdir(base_dir):

        path_to_file = os.path.join(base_dir, path)
        wb = load_workbook(filename=path_to_file)
        ws = wb.active

        export = UpdatePrice(
            worksheet=ws,
            margin=margin,
            original_margin=original_margin,
            currency_rate=current_course,
            rate_sell=rate_sell,
            vendor_code_column=vendor_code_column,
            valuta=valuta,
        )

        with open("epik_rate.json", "r", encoding="utf-8") as f:
            file = json.load(f)

        for item in file.get("rate"):
            if re.match(rf"{item}", path.split(".")[0]):
                logger.info(f"{item}:{file.get('rate').get(item)}")
                rate = file.get("rate").get(item)
                break

        for price in prices:
            export.updateEpik(rate=rate, from_price=os.getenv(price))

        wb.save(filename=path_to_file)


def rozetka(
    base_dir,
    prices,
    margin: int | float | str = 100,
    original_margin: int | float | str = 300,
    current_course: int | float = 39,
    rate_sell: int | float = 20,
    valuta: str = "USD",
    vendor_code_column: str = "E",
):
    wb = load_workbook(filename=base_dir)
    excel_tabs = wb.sheetnames

    def normalize_tab_id(tab_id: int | str | None) -> str:
        if isinstance(tab_id, (int, str)):
            return str(tab_id)
        return ""

    with open("rozetka_rate_id.json", "r", encoding="utf-8") as f:
        file_with_id = json.load(f)

    rates_with_id = file_with_id.get("rate")

    for tab in excel_tabs:
        tab_name: str = tab.split("|")[1]
        tab_id_row: str = tab.split("|")[0]
        tab_id = normalize_tab_id(tab_id_row)

        current_rate = rates_with_id.get(tab_id, {}).get("rate", 1)

        if current_rate == 1:
            logger.info(f" ****** Rozetka rate: {tab_name} not found *****")
            continue

        logger.info(
            f'{"*" * 5}    {tab_name} - ( {current_rate} )    {"*" * 5}')

        ws = wb[tab]

        export = UpdatePrice(
            worksheet=ws,
            margin=margin,
            original_margin=original_margin,
            currency_rate=current_course,
            rate_sell=rate_sell,
            vendor_code_column=vendor_code_column,
            valuta=valuta,
        )

        for price in prices:
            export.updateRozetka(
                rate=current_rate, from_price=os.getenv(price))

    for ws in wb.worksheets:
        ws.calculate_dimension()

    wb.save(filename=base_dir)


def manual(margin, original_margin, rate, rate_sell, currency_rate):
    export = UpdatePrice(currency_rate=currency_rate)
    rate = rate + (rate * 0.08)
    while True:
        price = float(input("Enter price: "))
        m = UpdatePrice._resolve_margin(margin, price)
        om = UpdatePrice._resolve_margin(original_margin, price)
        new_price = export.royalty(price + m, rate)
        old_price = export.royalty(new_price, rate_sell)
        or_new_price = export.royalty(price + om, rate)
        or_old_price = export.royalty(or_new_price, rate_sell)

        logger.warning('Manual price')
        print('new_price', new_price)
        print('old_price', old_price)
        print('or_new_price', or_new_price)
        print('or_old_price', or_old_price)


def main(marketplace):

    match marketplace:
        case "MANUAL":
            manual(
                margin=150,
                original_margin=480,
                rate=10,
                rate_sell=25,
                currency_rate=43.5,
            )

        case "PROM":
            prom(
                # prom_press, prom_korm, prom_incubator, prom_gaz, prom_grand, prom_secators, prom_welding, prom_sprayer
                export_file_path=prom_grand,
                prices=["GRAND_ELTOS"],  # GRAND_ELTOS | SPRAYER
                valuta="USD",  # UAH | USD
                current_course=45,
                margin='20%',  # 200 | 20%
                original_margin="50%",  # 530 | 50%
                rate_sell=25
            )

        case "EPICENTR":
            epicentr(
                base_dir=epic_grand,
                prices=["GRAND_ELTOS"],
                valuta="USD",
                current_course=45,
                margin=200,
                original_margin=530,
                rate_sell=25
            )
            # epic_press, epic_incubator, epic_korm, epic_gaz, epic_house_tech, epic_grand

        case "ROZETKA":
            rozetka(
                base_dir=rozetka_one_file,
                prices=["GRAND_ELTOS"],
                valuta="USD",
                current_course=45,
                margin='20%',
                original_margin='50%',
                rate_sell=25
            )
            # roz_press, roz_gaz, roz_incubator,roz_grand, rozetka_one_file, rozetka_secators, rozetka_saw, rozetka_welding

        case _:
            logger.info("You do not have any access to the code")


if __name__ == "__main__":
    # PRESS, KORM, GAZ, INCUBATOR, KITCHEN, GRAND_ELTOS

    MARKETPLACES = ["PROM", "ROZETKA", 'EPICENTR']
    # MARKETPLACES = ["PROM"]
    # MARKETPLACES = ["EPICENTR"]
    # MARKETPLACES = ["ROZETKA"]
    # MARKETPLACES = ["MANUAL"]
    #

    for MARKETPLACE in MARKETPLACES:
        logger.info("=" * 55)
        logger.info(f'{"=" * 20}    {MARKETPLACE}    {"=" * 20}')
        logger.info("=" * 55)
        main(MARKETPLACE)
