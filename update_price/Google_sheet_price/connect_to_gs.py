import logging
import re
import os
from pprint import pprint

import colorlog
import ezsheets
import dotenv
from openpyxl import load_workbook
import json
from dotenv import dotenv_values
import unicodedata

dotenv.load_dotenv()

handler = colorlog.StreamHandler()
handler.setFormatter(colorlog.ColoredFormatter(
    "%(log_color)s%(levelname)s:%(name)s:%(message)s",
    log_colors={
        'DEBUG': 'cyan',
        'INFO': 'green',
        'WARNING': 'yellow',
        'ERROR': 'red',
        'CRITICAL': 'bold_red',
    }
))

logger = colorlog.getLogger('my_logger')
logger.addHandler(handler)
logger.setLevel(logging.DEBUG)


class GPrice:

    def __init__(self,
                 sheet_id,
                 from_google,
                 path_file=None,
                 item_number_google_sheets='B',
                 title_google_sheets='C',
                 price_column_google_sheets='E',
                 recommended_retail_price_column_google_sheets='G',
                 amount_column_google_sheets='H',
                 available_column_google_sheets='I',
                 item_number_export_file='B',
                 price_column_export_file='H',
                 price_rrc_column_export_file='I',
                 available_column_export_file='G',
                 amount_column_export_file='F',

                 ):
        self.sheet_id = sheet_id
        self.from_google = from_google
        self.path_file = path_file
        self.item_number_google_sheets = item_number_google_sheets
        self.title_google_sheets = title_google_sheets
        self.price_column_google_sheets = price_column_google_sheets
        self.recommended_retail_price_column_google_sheets = recommended_retail_price_column_google_sheets
        self.available_column_google_sheets = available_column_google_sheets
        self.amount_column_google_sheets = amount_column_google_sheets
        self.item_number_export_file = item_number_export_file
        self.price_column_export_file = price_column_export_file
        self.price_rrc_column_export_file = price_rrc_column_export_file
        self.available_column_export_file = available_column_export_file
        self.amount_column_export_file = amount_column_export_file
        self.items_sheet_list = []
        self.items_excel_list = []

    @staticmethod
    def _col_letter_to_index(col: str) -> int:
        col = col.upper()
        result = 0
        for char in col:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    @staticmethod
    def _not_valid_item_number(vc):
        title_list = ['ФОТО', 'Артикул', 'Наименование товара', 'Основные параметры',
                      'Цена/USD', 'Цена/UAH', 'Количество в ящике', 'Наличие', 'Цена']

        if vc == '':
            return True
        elif vc is None:
            return True

        for name in title_list:
            if re.match(fr'^{name}', vc):
                return True

    @staticmethod
    def _replaceSpace(cell_value):
        return ' '.join(cell_value.split())

    @staticmethod
    def _valid_price(price):

        if isinstance(price, str):
            price = price.replace(',', '.')
            cleaned = re.sub(r'[^\d.\-]', '', price)
            try:
                return float(cleaned)
            except ValueError:
                return logger.error('Invalid price')
        try:
            return float(price)
        except Exception as ex:
            logger.error(f"Not valid rrc_price {price}, ERROR: {ex}")

    @staticmethod
    def _valid_availability_export_file(cell):

        if cell == '+':
            return 'TRUE'
        else:
            return 'FALSE'

    def connect_to_google(self, google_sheet_id):
        try:
            logger.info("Connecting to google sheets...")
            ss = ezsheets.Spreadsheet(google_sheet_id)
            sheet = ss.sheets[0]

            return sheet
        except Exception as ex:
            logger.error("Problem with connecting to google sheets...")
            print(ex)

    def _connect_to_excel(self):
        try:
            logger.info("Connecting to excel sheets...")
            ee = load_workbook(self.path_file)
            excel_sheet = ee.active
            return excel_sheet

        except Exception as ex:
            logger.error("Problem with connecting to excel sheets...")
            print(ex)

    def _get_data_from_excel_file_for_update_price(self):
        excel_sheet = self._connect_to_excel()

        excel_items = {}

        for i in range(1, excel_sheet.max_row + 1):
            item_number = excel_sheet[f'{self.item_number_export_file}{i}'].value

            if self._not_valid_item_number(item_number):
                continue

            price_excel = self._valid_price(
                excel_sheet[f"{self.price_column_export_file}{i}"].value)

            price_rrc_excel = self._valid_price(
                excel_sheet[f"{self.price_rrc_column_export_file}{i}"].value)

            if price_excel is None:
                continue

            amount = excel_sheet[f"{self.amount_column_export_file}{i}"].value

            availability_excel = self._valid_availability_export_file(
                excel_sheet[f"{self.available_column_export_file}{i}"].value)

            excel_items[item_number] = {
                'price': price_excel,
                'price_rrc': price_rrc_excel if price_rrc_excel else '0',
                'amount': amount,
                'availability': availability_excel
            }

        if not excel_items:
            logger.warning("The Excel file does not contain valid data.")
        return excel_items

    def _get_data_from_google_file_for_update_price(self):
        google_sheet = self.connect_to_google(self.from_google)
        all_rows = google_sheet.getRows()

        item_col = self._col_letter_to_index(self.item_number_export_file)
        price_col = self._col_letter_to_index(self.price_column_export_file)
        price_rrc_col = self._col_letter_to_index(self.price_rrc_column_export_file)
        amount_col = self._col_letter_to_index(self.amount_column_export_file)
        avail_col = self._col_letter_to_index(self.available_column_export_file)

        items = {}

        for row in all_rows:
            if len(row) <= item_col:
                continue

            item_number: str = row[item_col].strip()

            if self._not_valid_item_number(item_number):
                continue

            price = self._valid_price(row[price_col] if len(row) > price_col else '')
            price_rrc = self._valid_price(row[price_rrc_col] if len(row) > price_rrc_col else '')

            if price is None:
                continue

            amount = row[amount_col] if len(row) > amount_col else ''
            availability = self._valid_availability_export_file(
                row[avail_col] if len(row) > avail_col else '')

            items[item_number] = {
                'price': price,
                'price_rrc': price_rrc if price_rrc else '0',
                'amount': amount,
                'availability': availability
            }

        if not items:
            logger.warning("The Excel file does not contain valid data.")
        return items

    def _set_data_to_google_sheets_for_update_price(self, excel_items):

        if not excel_items:
            logger.warning("No data to update in Google Sheets.")
            return

        sheet = self.connect_to_google(self.sheet_id)
        all_rows = sheet.getRows()

        item_col = self._col_letter_to_index(self.item_number_google_sheets)
        title_col = self._col_letter_to_index(self.title_google_sheets)
        price_col = self._col_letter_to_index(self.price_column_google_sheets)
        rrp_col = self._col_letter_to_index(self.recommended_retail_price_column_google_sheets)
        amount_col = self._col_letter_to_index(self.amount_column_google_sheets)
        avail_col = self._col_letter_to_index(self.available_column_google_sheets)
        max_col = max(item_col, title_col, price_col, rrp_col, amount_col, avail_col)

        google_items = set()
        for row in all_rows:
            if len(row) > item_col:
                item_number = row[item_col]
                if not self._not_valid_item_number(item_number):
                    google_items.add(item_number)

        missing_items = {}
        for item_number, data in excel_items.items():
            if item_number not in google_items:
                missing_items[item_number] = data
                logger.info(f"Item {item_number} exists in Excel but not in Google Sheets")

        if missing_items:
            with open('missing_items.json', 'w', encoding='utf-8') as f:
                json.dump(missing_items, f, ensure_ascii=False, indent=4)
            logger.info(f"Missing items were saved to missing_items.json")

        for row in all_rows:
            while len(row) <= max_col:
                row.append('')

            item_number = row[item_col]

            if self._not_valid_item_number(item_number):
                continue

            row[title_col] = self._replaceSpace(row[title_col])

            if item_number in excel_items:
                data = excel_items[item_number]
                row[amount_col] = data['amount']
                row[price_col] = data['price']
                row[rrp_col] = data['price_rrc']
                row[avail_col] = data['availability']
                logger.info(f"Product data has been updated {item_number}.")
            else:
                row[avail_col] = "FALSE"
                logger.info(f"The item {item_number} was not found in Excel. Available set 'False'.")

        sheet.updateRows(all_rows)

    def _get_data_from_excel_file(self):
        logger.info("Getting data from an Excel file...")

        excel_sheet = self._connect_to_excel()

        excel_items = {}

        for i in range(2, excel_sheet.max_row + 1):
            item_number = excel_sheet[f'{self.item_number_export_file}{i}'].value

            if self._not_valid_item_number(item_number):
                continue

            rrp_price_excel = self._valid_price(
                excel_sheet[f"{self.price_rrc_column_export_file}{i}"].value)

            if rrp_price_excel is None:
                continue

            amount = excel_sheet[f"{self.amount_column_export_file}{i}"].value

            excel_items[item_number] = {
                'price': rrp_price_excel,
                'amount': amount
            }

        if not excel_items:
            logger.warning("The Excel file does not contain valid data.")
        return excel_items

    def _set_data_to_google_sheets(self, excel_items):

        if not excel_items:
            logger.warning("No data to update in Google Sheets.")
            return

        sheet = self.connect_to_google(self.sheet_id)
        all_rows = sheet.getRows()

        item_col = self._col_letter_to_index(self.item_number_google_sheets)
        amount_col = self._col_letter_to_index(self.amount_column_google_sheets)
        rrp_col = self._col_letter_to_index(self.recommended_retail_price_column_google_sheets)
        max_col = max(item_col, amount_col, rrp_col)

        for row in all_rows:
            while len(row) <= max_col:
                row.append('')

            item_number = row[item_col]

            if self._not_valid_item_number(item_number):
                continue

            if item_number in excel_items:
                data = excel_items[item_number]
                row[amount_col] = data['amount']
                row[rrp_col] = data['price']
                logger.info(f"Product data has been updated {item_number}.")
            else:
                row[rrp_col] = 0
                logger.info(f"The item {item_number} was not found in Excel. The price is set to 0.")

        sheet.updateRows(all_rows)

    def updatePrice(self):
        if self.path_file is None:
            items = self._get_data_from_google_file_for_update_price()
        else:
            items = self._get_data_from_excel_file_for_update_price()

        self._set_data_to_google_sheets_for_update_price(items)

    def updateOnlyRRP(self):

        excel_items = self._get_data_from_excel_file()

        self._set_data_to_google_sheets(excel_items)


def main():

    up = GPrice(
        sheet_id=dotenv_values()['GRAND_ELTOS'],
        from_google=dotenv_values()['PATH_GOOGLE_NEW_PRICE_RRP'],
        # path_file=dotenv_values()['PATH_FILE_EXCEL_PRICE_RRP'],
    )

    up.updatePrice()
    logger.debug('Completed')


if __name__ == '__main__':
    main()
