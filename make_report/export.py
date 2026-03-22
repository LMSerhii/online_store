from typing import Dict, List
import re
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed

from datetime import datetime

from config import PromExportConfig
from logger import logger
from status import StatusService
from clients import PromClient
from constans import patterns

from excel import ExcelConfig, ColumnAlignment


class ExportProm:
    def __init__(self, config: PromExportConfig, status_service: StatusService):
        self.config = config
        self.status_service = status_service
        self.api_client = PromClient(config)
        self.patterns = patterns
        self._compiled_patterns = {
            key: re.compile(pattern) for key, pattern in patterns.items()
        }

    def _format_date(self, date_str: str) -> str:
        try:

            dt = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f")
            return dt.date().isoformat()

        except Exception as e:
            logger.error(f"Failed to format date {date_str}: {e}")
            return date_str

    def _process_order_item(self, item: Dict, basic_info: Dict, is_first: bool) -> Dict:
        """Обробка окремого товару в замовленні"""

        try:
            return {
                "ID": basic_info["id"],
                "ПІБ": basic_info["client_name"],
                "Спосіб оплати": basic_info.get("payment_option_name", ""),
                "Кількість": item.get("quantity", ""),
                "Артикул": item.get("sku", ""),
                "Коментарі": basic_info["comments"] if is_first else "",
                "ТТН": basic_info["ttn"] or basic_info["barcode"],
                "Ціна продажу": basic_info["price"] if is_first else 0,
                "Ціна закупки": basic_info["total_purchase_price"] if is_first else 0,
                "Прибуток": "",
                "Ми >> Пе": basic_info["pe_column"] if is_first else "",
                "РС >> СЛ": basic_info["rs_column"] if is_first else "",
                "Денис >> СЛ'": basic_info["denis_column"] if is_first else "",
                "Мард >> СЛ": basic_info["marad_column"] if is_first else "",
                "Статус замовлення": basic_info["status"],
                "Товар": item.get("name", ""),
                "Спосіб замовлення": basic_info["order_type"],
                "Ціна доставки": (
                    basic_info["delivery_fields_delivery_cost"] if is_first else ""
                ),
                "Дата": self._format_date(basic_info["date"]),
            }

        except Exception as e:
            logger.error(f"Failed to process order item: {e}")
            return {}

    def _calculate_purchase_price(self, sku: str, quantity: int) -> float:
        """Розрахунок ціни закупки"""
        try:

            if sku is None:
                return 0

            separator = "||" if "||" in sku else "|"

            if not separator in sku:
                logger.warning(f"Couldn't find price in SKU: {sku}")
                return 0

            price = float(
                self._convert_price_text(sku.split(separator)[-1].lstrip("0")) or "0"
            )

            multiplier = float(self.config.current_course) if separator == "||" else 1

            return price * quantity * multiplier

        except Exception as e:
            logger.error(f"Failed to calculate purchase price for SKU {sku}: {e}")
            return 0

    def _priceFinderFromSku(self, sku: str) -> int:

        match = self._compiled_patterns["price_from_sku"].search(sku)

        return int(match.group(1)) if match else ""

    def _get_column_rs(
        self, payment_option_name: str | None, comments: str, price: float | int
    ):
        try:
            pattern = self._compiled_patterns["payment_option_name"]
            return (
                price
                if (
                    (pattern.search(payment_option_name) if payment_option_name else None)
                    or pattern.search(comments)
                )
                else ""
            )
        except Exception as e:
            logger.error(f"Failed to get rs column: {e}")
            return ""

    def _get_denis_column(self, comments: str, price: float) -> str:
        """Отримує значення для колонки denis з коментарів"""
        try:
            compiled = {
                "positive": self._compiled_patterns["denis_positive"],
                "negative": self._compiled_patterns["denis_negative"],
            }

            for pattern_type, pattern in compiled.items():
                if match := pattern.search(comments):
                    value = int(match.group(1))
                    if pattern_type == "negative":
                        return -value, price, value
                    else:
                        return value, value, 0
            return "", 0, 0
        except Exception as e:
            logger.error(f"Failed to get denis column: {e}")
            return "", 0, 0

    def _get_marad_column(self, comments: str) -> str:
        """Отримує значення для колонки marad з коментарів"""
        try:
            if match := self._compiled_patterns["marad"].search(comments):
                return int(match.group(1))
            return ""
        except Exception as e:
            logger.error(f"Failed to get marad column: {e}")
            return ""

    def _get_pe_column(self, commetns: str) -> str:
        """Отримує значення для колонки pe з коментарів"""
        try:
            if match := self._compiled_patterns["pe"].search(commetns):
                return int(match.group(1))
            return ""
        except Exception as e:
            logger.error(f"Failed to get pe column: {e}")
            return ""

    def _get_comments(self, labels) -> str:
        if not labels:
            return ""

        return ", ".join([label.get("name", "").replace(" ", "") for label in labels])

    def _convert_price_text(self, price_text: str) -> str:

        list_of_symbols = ["\xa0", " ", "₴"]

        for symbol in list_of_symbols:
            price_text = (
                price_text.replace(symbol, "") if symbol in price_text else price_text
            )

        price_text = price_text.replace(",", ".") if "," in price_text else price_text

        return price_text

    def _get_barcode_from_comments(self, comments):
        result = self._compiled_patterns["barcode"].search(comments)
        return result.group(0) if result else None

    def _fetch_page(self, page: int) -> List[Dict]:
        try:
            return self.api_client.get_orders(page) or []
        except Exception as e:
            logger.error(f"Failed to fetch page {page}: {e}")
            return []

    def _fetch_delivery(self, order_id, delivery_option_id):
        try:
            return self.api_client.get_delivery_data(
                id=order_id,
                delivery_option_id=delivery_option_id,
            )
        except Exception as e:
            logger.error(f"Failed to get delivery data for order {order_id}: {e}")
            return None, None

    def _fetch_status(self, order_id, barcode, delivery_option_id):
        try:
            if self.config.status == "true":
                return self.status_service.get_status(barcode=barcode)
            elif self.config.status == "false":
                return self.status_service.get_status_from_prom(
                    barcode=barcode,
                    delivery_option_id=delivery_option_id,
                )
        except Exception as e:
            logger.error(f"Failed to get status for order {order_id}: {e}")
        return ""

    def get_data(self) -> List[Dict]:
        """Основний метод для отримання всіх даних"""
        try:
            data_list = []
            pagination = self.api_client.get_pagination()

            if not pagination:
                logger.error("Failed to get pagination")
                return data_list

            # 1. Паралельне завантаження сторінок замовлень
            all_orders: List[Dict] = []
            pages = list(range(1, pagination + 1))
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(self._fetch_page, page): page for page in pages}
                for future in tqdm(as_completed(futures), total=len(futures), desc="Loading pages"):
                    result = future.result()
                    all_orders.extend(result)

            # 2. Паралельне отримання delivery_data для кожного замовлення
            delivery_map: Dict[int, tuple] = {}
            with ThreadPoolExecutor(max_workers=10) as executor:
                futures = {
                    executor.submit(
                        self._fetch_delivery,
                        order.get("id"),
                        order.get("delivery_option_id"),
                    ): order.get("id")
                    for order in all_orders
                }
                for future in tqdm(as_completed(futures), total=len(futures), desc="Loading delivery data"):
                    order_id = futures[future]
                    barcode, price = future.result()
                    delivery_map[order_id] = (barcode, price)

            # 3. Підготовка barcode (з fallback через коментарі) та delivery_option_id для статусів
            barcode_map: Dict[int, tuple] = {}
            for order in all_orders:
                order_id = order.get("id")
                comments = self._get_comments(order.get("labels", []))
                ttn = order.get("delivery_declaration_identifier", "")
                barcode, price = delivery_map.get(order_id, (None, None))
                barcode = barcode or self._get_barcode_from_comments(comments)
                effective_barcode = ttn or barcode
                barcode_map[order_id] = (effective_barcode, barcode, price)

            # 4. Паралельне отримання статусів
            status_map: Dict[int, str] = {}
            with ThreadPoolExecutor(max_workers=7) as executor:
                futures = {
                    executor.submit(
                        self._fetch_status,
                        order.get("id"),
                        barcode_map[order.get("id")][0],
                        order.get("delivery_option_id"),
                    ): order.get("id")
                    for order in all_orders
                    if barcode_map.get(order.get("id"), (None,))[0]
                }
                for future in tqdm(as_completed(futures), total=len(futures), desc="Loading statuses"):
                    order_id = futures[future]
                    status_map[order_id] = future.result() or ""

            # 5. Збірка data_list
            for order in all_orders:
                order_id = order.get("id", "")
                comments = self._get_comments(order.get("labels", []))
                _, barcode, price = barcode_map.get(order_id, (None, None, None))

                total_purchase_price = sum(
                    self._calculate_purchase_price(item["sku"], item["quantity"])
                    for item in order.get("added_items", [])
                )

                basic_info = {
                    "id": order_id,
                    "date": order.get("created", ""),
                    "order_type": order.get("type", ""),
                    "client_name": f"{order.get('client_first_name', '')} {order.get('client_last_name', '')}",
                    "comments": comments,
                    "payment_option_name": order.get("payment_option_name", ""),
                    "delivery_option_id": order.get("delivery_option_id"),
                    "delivery_fields_delivery_cost": order.get(
                        "delivery_fields_delivery_cost", "-"
                    ),
                    "cart_total_price": order.get("price_text", ""),
                    "ttn": order.get("delivery_declaration_identifier", ""),
                }

                denis_column, denis_price, denis_purchase = self._get_denis_column(
                    basic_info["comments"], price
                )

                marad_column = self._get_marad_column(basic_info["comments"])
                pe_column = self._get_pe_column(basic_info["comments"])
                price_from_sku = self._priceFinderFromSku(basic_info["comments"])

                price = next(
                    filter(None, [denis_price, marad_column, price_from_sku, price]),
                    "",
                )

                if pe_column:
                    total_purchase_price += pe_column

                if denis_purchase:
                    total_purchase_price = denis_purchase

                rs_column = self._get_column_rs(
                    basic_info["payment_option_name"], basic_info["comments"], price
                )

                basic_info.update(
                    {
                        "barcode": barcode,
                        "price": price,
                        "rs_column": rs_column,
                        "denis_column": denis_column,
                        "marad_column": marad_column,
                        "pe_column": pe_column,
                        "total_purchase_price": total_purchase_price,
                        "status": status_map.get(order_id, ""),
                    }
                )

                for idx, item in enumerate(order["added_items"]):
                    order_data = self._process_order_item(
                        item, basic_info, is_first=(idx == 0)
                    )
                    if order_data:
                        data_list.append(order_data)

            return data_list

        except Exception as e:
            logger.error(f"Failed to get data: {e}")
            return []

    def export_to_excel(self, data: List[Dict], filename: str):
        try:
            import pandas as pd
            from openpyxl.styles import Alignment, PatternFill
            from openpyxl.formatting.rule import CellIsRule

            df = pd.DataFrame(data)

            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Sheet1")
                ws = writer.sheets["Sheet1"]

                # Автоматична ширина
                for column in ws.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column)
                    col_letter = column[0].column_letter
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Застосування вирівнювання
                alignment_map = {
                    ColumnAlignment.CENTER: Alignment(
                        horizontal="center", vertical="center"
                    ),
                    ColumnAlignment.LEFT: Alignment(
                        horizontal="left", vertical="center"
                    ),
                    ColumnAlignment.RIGHT: Alignment(
                        horizontal="right", vertical="center"
                    ),
                }

                for alignment_type in ColumnAlignment:
                    columns = ExcelConfig.get_columns_by_alignment(alignment_type)
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.column_letter in columns:
                                cell.alignment = alignment_map[alignment_type]

                ws.freeze_panes = "A1"

                # ws.conditional_formatting.add(
                #     "D2:D1000",
                #     CellIsRule(
                #         operator="greaterThan",
                #         formula=["0"],
                #         stopIfTrue=True,
                #         fill=PatternFill(
                #             start_color="c6efce", end_color="c6efce", fill_type="solid"
                #         ),
                #     ),
                # )

            logger.info(f"Data exported successfully to {filename}")

        except Exception as e:
            logger.error(f"Failed to export data to Excel: {e}")
